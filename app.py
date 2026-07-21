from flask import Flask, render_template, request, jsonify, flash, redirect, url_for
from flask_socketio import SocketIO
import os
from werkzeug.utils import secure_filename
import openpyxl
from datetime import datetime, date
from decimal import Decimal, ROUND_HALF_UP

import math
import requests
from db import (
    load_env_from_dotenv,
    get_hana_connection,
    insert_archivoplano,
    upsert_archivoplano,
    truncate_archivoplano,
)
import socket
from werkzeug.middleware.proxy_fix import ProxyFix


# Configuración DNS para SAP BTP - SOLO UNA VEZ
def configure_dns_for_sap_btp():
    if "VCAP_APPLICATION" in os.environ:
        print("Configurando DNS para SAP BTP...")
        original_getaddrinfo = socket.getaddrinfo

        def patched_getaddrinfo(host, port, family=0, type=0, proto=0, flags=0):
            try:
                return original_getaddrinfo(host, port, family, type, proto, flags)
            except socket.gaierror as e:
                print(f"DNS resolution failed for {host}, trying IP fallback...")
                if host == "telcl-dev-db-cap-telcl-srv.cfapps.us10.hana.ondemand.com":
                    return [
                        (
                            socket.AF_INET,
                            socket.SOCK_STREAM,
                            6,
                            "",
                            ("52.23.1.211", port),
                        )
                    ]
                raise e

        socket.getaddrinfo = patched_getaddrinfo


# Llamar configuración DNS UNA SOLA VEZ
configure_dns_for_sap_btp()
load_env_from_dotenv()

app = Flask(__name__)
# Solución BTP Proxy: Respeta los headers X-Forwarded de Cloud Foundry para el enrutamiento y 404
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

app.secret_key = "Hitss_REP4_Flask_2025"
socketio = SocketIO(app, cors_allowed_origins="*")

# Configuración de la aplicación
app.config["UPLOAD_FOLDER"] = "uploads"
app.config["BOT_FOLDER"] = "bot_files"

if not os.path.exists(app.config["BOT_FOLDER"]):
    os.makedirs(app.config["BOT_FOLDER"])

app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024  # 200 MB
ALLOWED_EXTENSIONS = {"xlsx", "xls", "xlsb"}

# Constantes
BATCH_SIZE = int(os.getenv("BATCH_SIZE", "100"))
MAX_CONCURRENCY = 500

HTTP_TIMEOUT = float(os.getenv("HTTP_TIMEOUT", "30"))
if not os.path.exists(app.config["UPLOAD_FOLDER"]):
    os.makedirs(app.config["UPLOAD_FOLDER"])


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def encontrar_fila_encabezados(sheet):
    """Encuentra la fila de encabezados del ArchivoPlano buscando 'RPU' en col 1."""
    # En el ArchivoPlano la fila de encabezados tiene 'RPU' en la primera columna
    for row_idx in range(1, 50):
        val = sheet.cell(row=row_idx, column=1).value
        if val is not None and str(val).strip().upper() == "RPU":
            return row_idx
    return 1  # fallback


def leer_periodo_del_archivo(sheet):
    """
    Lee el valor de 'Periodo' desde las primeras filas del ArchivoPlano.
    El archivo tiene en col1='Periodo', col2='YYYY-MM' (ej. '2024-08').
    Retorna (mes, anio) como strings, o (None, None) si no se encuentra.
    """
    for row_idx in range(1, 20):
        val = sheet.cell(row=row_idx, column=1).value
        if val is not None and str(val).strip().lower() == "periodo":
            periodo = sheet.cell(row=row_idx, column=2).value
            if periodo:
                partes = str(periodo).strip().split("-")
                if len(partes) == 2:
                    return partes[1], partes[0]  # (mes, anio)
    return None, None


def procesar_excel(filepath, fecha_facturacion=None):
    """
    Procesa el ArchivoPlano (formato M249).
    - Lee el Periodo directamente del archivo.
    - Extrae solo las columnas de interés por índice (0-based).
    - Usa read_only=True + iteración nativa por filas para máxima velocidad.
    """

    # Columnas de interés: nombre → índice 0-based (número de columna - 1)
    COLS_INTERES = {
        "RPU": 0,
        "Periodo": 1,
        "TipoMov": 2,
        "DigVer": 3,
        "Importe total": 11,
        "Importe energía total": 12,
        "Importe IVA": 13,
        "Motivo": 18,
        "Importe": 19,
        "Consumo": 80,
        "Demanda": 81,
        "Reactivos": 82,
        "Nombre": 242,
        "Dirección": 243,
        "Ciudad": 244,
        "Estado": 245,
        "RFC": 246,
        "Colonia": 247,
        "Calle 1": 248,
        "Calle 2": 249,
        "IdEmpresa": 250,
        "Numero": 97,
        "Cuenta": 230,
        "CargaContratada": 234,
    }
    col_indices = list(COLS_INTERES.values())

    try:
        is_xlsb = filepath.lower().endswith(".xlsb")
        if is_xlsb:
            import pyxlsb

            workbook_xlsb = pyxlsb.open_workbook(filepath)
            sheet_names = workbook_xlsb.sheets
            workbook = None
        else:
            workbook = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            sheet_names = workbook.sheetnames
            workbook_xlsb = None

        num_sheets = len(sheet_names)
        hojas_procesadas = []
        mes, anio = "00", "0000"

        for sheet_name in sheet_names:
            encabezados = list(COLS_INTERES.keys())
            filas_validas = []

            fila_actual = 0
            fila_inicio = None
            header_found = False

            if is_xlsb:
                sheet_xlsb = workbook_xlsb.get_sheet(sheet_name)
                row_iterator = ([c.v for c in r] for r in sheet_xlsb.rows())
            else:
                sheet = workbook[sheet_name]
                row_iterator = sheet.iter_rows(values_only=True)

            for row_cells in row_iterator:
                fila_actual += 1

                # ── Buscar Periodo (primeras 20 filas) ──────────────────────
                if fila_actual <= 20:
                    if (
                        row_cells
                        and str(row_cells[0] or "").strip().lower() == "periodo"
                    ):
                        periodo_val = row_cells[1] if len(row_cells) > 1 else None
                        if periodo_val:
                            partes = str(periodo_val).strip().split("-")
                            if len(partes) == 2:
                                anio, mes = partes[0], partes[1]

                # ── Detectar fila de encabezados (col1 == 'RPU') ────────────
                if not header_found:
                    if row_cells and str(row_cells[0] or "").strip().upper() == "RPU":
                        fila_inicio = fila_actual
                        header_found = True
                    continue

                # ── A partir de aquí son filas de datos ─────────────────────
                if not row_cells:
                    continue

                rpu_val = row_cells[0] if len(row_cells) > 0 else None
                if rpu_val is None or str(rpu_val).strip() == "":
                    continue

                rpu_str = str(rpu_val).strip().upper()
                if rpu_str.startswith("SUBTOTAL") or rpu_str.startswith("TOTAL"):
                    continue

                # Extraer solo los índices de interés — valores crudos
                fila_datos = [
                    row_cells[idx] if idx < len(row_cells) else None
                    for idx in col_indices
                ]
                # Agregar nombre de pestaña al final (campo IVATYPE)
                fila_datos.append(sheet_name)

                filas_validas.append(fila_datos)

            total_filas_reales = len(filas_validas)

            hojas_procesadas.append(
                {
                    "nombre": sheet_name,
                    "encabezados": encabezados,
                    "datos": filas_validas,
                    "datos_preview": filas_validas[:40],
                    "fila_inicio": fila_inicio or 0,
                    "total_filas": total_filas_reales,
                    "filas_eliminadas": 0,
                    "periodo": f"{anio}-{mes.zfill(2)}",
                    "headerperiod": f"{anio}{mes.zfill(2)}",
                }
            )

            if is_xlsb:
                sheet_xlsb.close()

        if is_xlsb:
            workbook_xlsb.close()
        else:
            workbook.close()

        return {
            "num_hojas": num_sheets,
            "nombres_hojas": sheet_names,
            "hojas": hojas_procesadas,
            "fecha_facturacion": f"{mes}/{anio}",
            "nombre_archivo": os.path.basename(filepath),
        }

    except Exception as e:
        raise Exception(f"Error al procesar el archivo Excel: {str(e)}")


def formatear_fecha(valor):
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%Y-%m-%d")  # solo fecha sin hora
    return valor


def formatear_valor(valor):
    # Fechas → se manejan aparte
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%Y-%m-%d")

    # Números → los pasamos a Decimal para evitar flotantes
    if isinstance(valor, (int, float)):
        dec = Decimal(str(valor)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return f"{dec:,}"  # añade comas como separadores de miles

    return valor


def normalize_tarifa(tarifa_value):
    if tarifa_value is None:
        return "00"

    str_tarifa = str(tarifa_value).strip()

    # Caso 1: vacío
    if str_tarifa == "":
        return "00"

    # Caso 2: 1 dígito numérico
    if len(str_tarifa) == 1 and str_tarifa.isdigit():
        return f"0{str_tarifa}"

    # Caso 3: ya tiene 2 caracteres
    if len(str_tarifa) == 2:
        return str_tarifa

    # Caso 4: más de 2 caracteres → tomar primeros 2
    if len(str_tarifa) > 2:
        return str_tarifa[:2]

    # Caso por defecto
    return "00"


def a_decimal(valor):
    """Convierte cualquier valor a Decimal, limpiando comas si existen"""
    return Decimal(str(valor).replace(",", ""))


def mapear_registro(fila, headerperiod=""):
    """
    Mapea una fila del ArchivoPlano a un dict para insertar en HANA.
    Aplica el tipo exacto que espera cada campo de la BD:
      - String  → str, nunca float
      - Decimal → float con 2 decimales
    """

    # print("----- VALORES RAW DESDE EXCEL -----")
    # print("ENERGYAMOUNT RAW:", fila[5], repr(fila[5]), type(fila[5]))
    # print("IVA RAW:", fila[6], repr(fila[6]), type(fila[6]))
    # print("-----------------------------------")

    def to_str(v):
        """String limpio — nunca agrega decimales a enteros."""
        if v is None:
            return ""
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()

    def to_decimal(v):
        try:
            if v in (None, "", "None"):
                return Decimal("0")
            return Decimal(str(v))
        except:
            return Decimal("0")

    return {
        # Keys (String)
        "RPU": to_str(fila[0]),
        "HEADERPERIOD": headerperiod,
        "PERIOD": to_str(fila[1]),
        "MOVTYPE": to_str(fila[2]),
        "CHECKDIG": to_str(fila[3]),
        # Decimals
        "TOTALAMOUNT": to_decimal(fila[4]),
        "ENERGYAMOUNT": to_decimal(fila[5]),
        "IVA": to_decimal(fila[6]),
        # String
        "REASON": to_str(fila[7]),
        # Decimals
        "AMOUNT": to_decimal(fila[8]),
        "CONSUMPTION": to_decimal(fila[9]),
        "DEMAND": to_decimal(fila[10]),
        "REACTIVEPOWER": to_decimal(fila[11]),
        # Strings
        "CUSTOMERNAME": to_str(fila[12]),
        "ADDRESS": to_str(fila[13]),
        "CITY": to_str(fila[14]),
        "STATE": to_str(fila[15]),
        "RFC": to_str(fila[16]),
        "NEIGHBORHOOD": to_str(fila[17]),
        "STREET1": to_str(fila[18]),
        "STREET2": to_str(fila[19]),
        "COMPANYID": to_str(fila[20]),
        "METER": to_str(fila[21]),
        "ACCOUNT": to_str(fila[22]),
        "CONTRACTEDDEMAND": to_decimal(fila[23]),
        # Campo calculado — IVATYPE ahora corre del índice 21 al 24
        "IVATYPE": to_str(fila[24]) if len(fila) > 24 else "",
    }


def procesar_hoja_db(hoja, session_id, modo):
    """
    Procesa una hoja e inserta sus registros en HANA en batches.
    Completamente sincrono: compatible con gunicorn eventlet sin conflictos.
    """
    registros = hoja.get("datos", [])
    hoja_nombre = hoja.get("nombre")
    total = len(registros)
    errores = 0
    exitos = 0
    registros_procesados = 0
    conn = get_hana_connection()
    try:
        for i in range(0, total, BATCH_SIZE):
            batch = registros[i : i + BATCH_SIZE]
            entities = []
            headerperiod = hoja.get("headerperiod", "")
            for r in batch:
                entity = mapear_registro(r, headerperiod)

                # print("---- REGISTRO A INSERTAR ----")
                # print("ENERGYAMOUNT:", entity["ENERGYAMOUNT"], type(entity["ENERGYAMOUNT"]))
                # print("IVA:", entity["IVA"], type(entity["IVA"]))
                # print("-----------------------------")

                entities.append(entity)
            if modo == "upsert":
                result = upsert_archivoplano(conn, entities)
            else:
                result = insert_archivoplano(conn, entities)
            processed = (
                result.get("updated", 0)
                + result.get("inserted", 0)
                + result.get("failed", 0)
            )
            registros_procesados += processed
            exitos += result.get("updated", 0) + result.get("inserted", 0)
            errores += result.get("failed", 0)
            if result.get("errors"):
                print(f"Errores en lote {i//BATCH_SIZE+1}: {len(result['errors'])}")
                for e in result["errors"][:5]:
                    print(e)
            print(
                f"Lote {i//BATCH_SIZE+1} hoja {hoja_nombre}: "
                f"ok={exitos}, errores={errores}, procesados={registros_procesados}/{total}"
            )
            progress = (registros_procesados / total) * 100 if total else 100
            socketio.emit(
                "progress_update",
                {
                    "current": registros_procesados,
                    "total": total,
                    "progress": round(progress, 2),
                },
                room=session_id,
            )
    finally:
        conn.close()
    return {"hoja": hoja_nombre, "total": total, "exitos": exitos, "errores": errores}


# ***************************************************************************************************************


@app.route("/enviar_hoja", methods=["POST"])
def enviar_hoja():
    """
    Recibe UNA sola hoja y la inserta en HANA.
    El frontend envia hoja por hoja para no superar el timeout de 60s del GoRouter de CF.
    El frontend llama a /borrar_datos una sola vez ANTES de la primera hoja.
    """
    try:
        data = request.get_json(force=True)
        hoja = data.get("hoja")
        if not hoja:
            return jsonify(
                {"success": False, "error": "Falta el campo 'hoja' en el body"}
            )
        session_id = request.args.get("session_id")
        modo = request.args.get("mode", "insert")

        resultado = procesar_hoja_db(hoja, session_id, modo)
        return jsonify({"success": True, "resultado": resultado})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/enviar_datos", methods=["POST"])
def enviar_datos():
    """Endpoint legacy - mantiene compatibilidad. Para BTP usar /enviar_hoja."""
    try:
        data = request.get_json(force=True)
        hojas = data.get("hojas", [])
        session_id = request.args.get("session_id")
        modo = request.args.get("mode", "insert")

        resultados = [procesar_hoja_db(h, session_id, modo) for h in hojas]
        return jsonify({"success": True, "resultados": resultados})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# Agregar manejo de conexiones SocketIO
@socketio.on("connect")
def handle_connect():
    print("Cliente conectado:", request.sid)


@socketio.on("disconnect")
def handle_disconnect():
    print("Cliente desconectado:", request.sid)


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        # Verificar si se envió el archivo
        if "excelFile" not in request.files:
            flash("No se encontró el archivo en la solicitud", "error")
            return redirect(request.url)

        file = request.files["excelFile"]
        fecha_facturacion = request.form.get("fechaFacturacion")

        # Validar que se haya seleccionado un archivo
        if file.filename == "":
            flash("No se seleccionó ningún archivo", "error")
            return redirect(request.url)

        # Validar extensión del archivo
        if file and allowed_file(file.filename):
            try:
                pass

                # Guardar el archivo
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
                file.save(filepath)

                # Procesar el archivo Excel (el periodo se lee del propio archivo)
                resultado = procesar_excel(filepath)

                # ✅ Borrar el archivo después de procesarlo
                try:
                    os.remove(filepath)
                except Exception:
                    pass  # Si falla el borrado, no es crítico

                # Renderizar la plantilla con los resultados
                return render_template("index.html", resultado=resultado)

            except Exception as e:
                flash(f"Error al procesar el archivo: {str(e)}", "error")
                return redirect(request.url)
        else:
            flash(
                "Tipo de archivo no permitido. Solo se aceptan archivos Excel (.xlsx, .xls)",
                "error",
            )
            return redirect(request.url)

    # Método GET - mostrar formulario vacío
    return render_template("index.html")


def delete_all_data():
    try:
        conn = get_hana_connection()
        truncate_archivoplano(conn)
        conn.close()
        return {"success": True, "deleted_count": 0, "message": "Tabla vaciada"}
    except Exception as e:
        return {"success": False, "deleted_count": 0, "message": str(e)}


@app.route("/borrar_datos", methods=["POST", "GET"])
def borrar_datos():
    try:
        print("=== Iniciando borrar_datos ===")
        result = delete_all_data()
        print(f"Resultado: {result}")
        return jsonify(result)
    except Exception as e:
        print(f"Error en borrar_datos: {e}")
        return (
            jsonify(
                {
                    "success": False,
                    "deleted_count": 0,
                    "message": f"Error interno: {str(e)}",
                }
            ),
            500,
        )


@app.route("/test-delete")
def test_delete():
    result = delete_all_data()
    return jsonify({"test_result": result, "timestamp": datetime.now().isoformat()})


@app.route("/test-connection")
def test_connection():
    results = {}

    # Test con hostname
    try:
        target = "https://telcl-dev-db-cap-telcl-srv.cfapps.us10.hana.ondemand.com"
        resp = requests.get(target, timeout=10)
        results["hostname_test"] = {"success": True, "status": resp.status_code}
    except Exception as e:
        results["hostname_test"] = {"success": False, "error": str(e)}

    # Test con IP directa
    try:
        resp = requests.get(
            "https://52.23.1.211",
            headers={
                "Host": "telcl-dev-db-cap-telcl-srv.cfapps.us10.hana.ondemand.com"
            },
            timeout=10,
            verify=False,
        )
        results["ip_test"] = {"success": True, "status": resp.status_code}
    except Exception as e:
        results["ip_test"] = {"success": False, "error": str(e)}

    # Test delete
    try:
        delete_result = delete_all_data()
        results["delete_test"] = delete_result
    except Exception as e:
        results["delete_test"] = {"success": False, "error": str(e)}

    return jsonify(results)


@app.route("/debug-connectivity")
def debug_connectivity():
    hostname = "telcl-dev-db-cap-telcl-srv.cfapps.us10.hana.ondemand.com"
    results = {}

    # Test DNS
    try:
        ip = socket.gethostbyname(hostname)
        results["dns_resolution"] = {"success": True, "ip": ip}
    except Exception as e:
        results["dns_resolution"] = {"success": False, "error": str(e)}

    # Test port connectivity
    try:
        sock = socket.create_connection((hostname, 443), timeout=10)
        sock.close()
        results["port_connectivity"] = {"success": True}
    except Exception as e:
        results["port_connectivity"] = {"success": False, "error": str(e)}

    # Test HTTP
    try:
        session = requests.Session()
        resp = session.get(f"https://{hostname}", timeout=30)
        results["http_test"] = {"success": True, "status": resp.status_code}
    except Exception as e:
        results["http_test"] = {"success": False, "error": str(e)}

    return jsonify(results)


@app.route("/limpiar_y_redirigir")
def limpiar_y_redirigir():
    return redirect(url_for("index"))


@app.route("/health")
def health_check():
    return jsonify({"status": "healthy", "message": "Service is running"})


############################## Recibir archivo bot #####################################


@app.route("/bot_upload", methods=["POST"])
def bot_upload():
    """
    Endpoint para recibir archivos desde el bot de SAP Process Automation.
    Recibe el archivo via multipart/form-data, lo guarda en bot_files/,
    lo procesa e inserta en HANA, y lo borra al terminar.
    No requiere intervención del usuario.

    Validaciones:
        - Campo 'file' presente en la request
        - Nombre de archivo no vacío
        - Extensión permitida (xlsx, xls, xlsb)
        - Tamaño mínimo del archivo (no vacío/corrupto)
        - Hojas requeridas presentes (T16, T8N, T8S)
        - Al menos una hoja con datos
        - Periodo legible en el archivo
    """
    filepath = None
    try:
        # ── 1. Validar que se envió el campo 'file' ──────────────────────────
        if "file" not in request.files:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "No se encontró el campo 'file' en la request. Verifica que el bot esté enviando el archivo correctamente.",
                    }
                ),
                400,
            )

        file = request.files["file"]

        # ── 2. Validar nombre de archivo ─────────────────────────────────────
        if file.filename == "":
            return (
                jsonify(
                    {"success": False, "error": "El archivo recibido no tiene nombre."}
                ),
                400,
            )

        # ── 3. Validar extensión ─────────────────────────────────────────────
        if not allowed_file(file.filename):
            return (
                jsonify(
                    {
                        "success": False,
                        "error": f"Extensión no permitida: '{file.filename.rsplit('.', 1)[-1]}'. Solo se aceptan: xlsx, xls, xlsb.",
                    }
                ),
                400,
            )

        # ── 4. Validar tamaño mínimo (archivo no vacío/corrupto) ────────────
        file.seek(0, 2)  # ir al final del stream
        file_size = file.tell()
        file.seek(0)  # regresar al inicio antes de guardarlo

        if file_size < 1024:  # menos de 1KB → sospechoso
            return (
                jsonify(
                    {
                        "success": False,
                        "error": f"El archivo es demasiado pequeño ({file_size} bytes). Puede estar vacío o corrupto.",
                    }
                ),
                400,
            )

        # ── 5. Guardar en bot_files/ ─────────────────────────────────────────
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config["BOT_FOLDER"], filename)
        file.save(filepath)
        print(f"[bot_upload] Archivo guardado: {filepath} ({file_size} bytes)")

        # ── 6. Parsear el Excel ──────────────────────────────────────────────
        try:
            resultado = procesar_excel(filepath)
        except Exception as e:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": f"Error al leer el archivo. Puede estar malformado o no ser un Excel válido. Detalle: {str(e)}",
                    }
                ),
                422,
            )

        # ── 7. Validar que el periodo sea legible ────────────────────────────
        if resultado.get("fecha_facturacion") == "00/0000":
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "No se pudo leer el periodo del archivo. Verifica que el bot procesó correctamente el archivo original.",
                    }
                ),
                422,
            )

        # ── 8. Validar hojas requeridas ──────────────────────────────────────
        hojas_encontradas = {h["nombre"] for h in resultado["hojas"]}
        hojas_requeridas = {"T16", "T8N", "T8S"}
        hojas_faltantes = hojas_requeridas - hojas_encontradas

        if hojas_faltantes:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": f"El archivo no contiene las hojas requeridas: {sorted(hojas_faltantes)}. "
                        f"Hojas encontradas: {sorted(hojas_encontradas)}.",
                    }
                ),
                422,
            )

        # ── 9. Validar que al menos una hoja tenga datos ─────────────────────
        hojas_vacias = [
            h["nombre"] for h in resultado["hojas"] if h["total_filas"] == 0
        ]

        if len(hojas_vacias) == len(resultado["hojas"]):
            return (
                jsonify(
                    {
                        "success": False,
                        "error": f"Todas las hojas están vacías: {hojas_vacias}. El archivo no contiene datos para procesar.",
                    }
                ),
                422,
            )

        if hojas_vacias:
            # Advertencia pero no bloqueante — algunas hojas pueden estar vacías
            print(f"[bot_upload] Advertencia: hojas sin datos: {hojas_vacias}")

        # ── 10. Truncar tabla e insertar hoja por hoja ───────────────────────
        delete_result = delete_all_data()
        if not delete_result.get("success"):
            return (
                jsonify(
                    {
                        "success": False,
                        "error": f"Error al limpiar la tabla antes de insertar: {delete_result.get('message')}",
                    }
                ),
                500,
            )

        resultados_hojas = []
        for hoja in resultado["hojas"]:
            res = procesar_hoja_db(hoja, session_id=None, modo="insert")
            resultados_hojas.append(res)

        # ── 11. Respuesta enriquecida para el bot ────────────────────────────
        total_registros = sum(r["total"] for r in resultados_hojas)
        total_exitos = sum(r["exitos"] for r in resultados_hojas)
        total_errores = sum(r["errores"] for r in resultados_hojas)

        print(
            f"[bot_upload] Carga completada: {total_exitos}/{total_registros} registros exitosos"
        )

        return jsonify(
            {
                "success": True,
                "archivo": filename,
                "periodo": resultado.get("fecha_facturacion"),
                "resumen": {
                    "total_registros": total_registros,
                    "total_exitos": total_exitos,
                    "total_errores": total_errores,
                },
                "hojas": [
                    {
                        "nombre": r["hoja"],
                        "total": r["total"],
                        "exitos": r["exitos"],
                        "errores": r["errores"],
                    }
                    for r in resultados_hojas
                ],
                "advertencias": (
                    [f"Hoja '{h}' no contiene datos" for h in hojas_vacias]
                    if hojas_vacias
                    else []
                ),
            }
        )

    except Exception as e:
        print(f"[bot_upload] Error inesperado: {e}")
        return (
            jsonify({"success": False, "error": f"Error interno inesperado: {str(e)}"}),
            500,
        )

    finally:
        # Borrar el archivo siempre, haya error o no
        try:
            if filepath and os.path.exists(filepath):
                os.remove(filepath)
                print(f"[bot_upload] Archivo borrado: {filepath}")
        except Exception:
            pass


# SocketIO handlers
@socketio.on("connect")
def handle_connect():
    print("Cliente conectado:", request.sid)


@socketio.on("disconnect")
def handle_disconnect():
    print("Cliente desconectado:", request.sid)


if __name__ == "__main__":
    socketio.run(app, debug=True, allow_unsafe_werkzeug=True)

# This is a comment only for merge purpose
# This is a comment only for merge purpose too

# Only for merge
